"""
E-ComSight — Export Service (CSV, Excel, PDF)
"""
import io
import json
import logging
from datetime import datetime
from typing import List
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT

logger = logging.getLogger(__name__)

SENTIMENT_VI = {
    "positive": "Tích cực",
    "neutral": "Trung lập",
    "negative": "Tiêu cực",
}
ASPECT_VI = {
    "product": "Chất lượng SP",
    "shipping": "Vận chuyển",
    "service": "Dịch vụ bán",
    "price": "Giá cả",
}
URGENCY_VI = {
    "critical": "Nghiêm trọng",
    "high": "Cao",
    "medium": "Trung bình",
    "low": "Thấp",
}


def reviews_to_dataframe(reviews: list) -> pd.DataFrame:
    """Convert list of Review objects to DataFrame"""
    rows = []
    for r in reviews:
        rows.append({
            "ID": r.id,
            "Nền tảng": r.platform.capitalize() if r.platform else "",
            "Tên sản phẩm": r.product_name or "",
            "Nội dung review": r.comment or "",
            "Số sao": r.rating_star or 0,
            "Cảm xúc": SENTIMENT_VI.get(r.sentiment_label, r.sentiment_label or ""),
            "Độ tin cậy": f"{round((r.sentiment_score or 0) * 100, 1)}%",
            "Khía cạnh": ASPECT_VI.get(r.aspect_label, r.aspect_label or ""),
            "Mức độ khẩn cấp": URGENCY_VI.get(r.urgency_label, r.urgency_label or ""),
            "Tác giả": r.author_username or "",
            "Ngày review": r.review_date.strftime("%d/%m/%Y") if r.review_date else "",
            "Đã gán nhãn": "Có" if r.is_labeled else "Tự động",
        })
    return pd.DataFrame(rows)


def export_csv(reviews: list) -> bytes:
    """Export reviews to CSV"""
    df = reviews_to_dataframe(reviews)
    output = io.StringIO()
    df.to_csv(output, index=False, encoding="utf-8-sig")
    return output.getvalue().encode("utf-8-sig")


def export_excel(reviews: list) -> bytes:
    """Export reviews to Excel with formatting"""
    df = reviews_to_dataframe(reviews)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Reviews", index=False)
        workbook = writer.book
        worksheet = writer.sheets["Reviews"]

        # Header formatting
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        header_fill = PatternFill("solid", fgColor="1E40AF")
        header_font = Font(bold=True, color="FFFFFF", size=11)

        for col_idx, col in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Sentiment color coding
        sentiment_colors = {
            "Tích cực": "DCFCE7",  # green-100
            "Trung lập": "F1F5F9",  # slate-100
            "Tiêu cực": "FEE2E2",  # red-100
        }
        sentiment_col = df.columns.get_loc("Cảm xúc") + 1

        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=sentiment_col)
            color = sentiment_colors.get(cell.value, "FFFFFF")
            cell.fill = PatternFill("solid", fgColor=color)

        # Column widths
        col_widths = {"Nội dung review": 60, "Tên sản phẩm": 30}
        for col_idx, col in enumerate(df.columns, 1):
            width = col_widths.get(col, max(len(col) + 4, 12))
            worksheet.column_dimensions[worksheet.cell(1, col_idx).column_letter].width = width

        # Summary sheet
        summary_data = {
            "Tổng reviews": len(reviews),
            "Tích cực": df["Cảm xúc"].value_counts().get("Tích cực", 0),
            "Trung lập": df["Cảm xúc"].value_counts().get("Trung lập", 0),
            "Tiêu cực": df["Cảm xúc"].value_counts().get("Tiêu cực", 0),
            "Shopee": df["Nền tảng"].value_counts().get("Shopee", 0),
            "TikTok": df["Nền tảng"].value_counts().get("Tiktok", 0),
        }
        df_summary = pd.DataFrame([{"Chỉ số": k, "Giá trị": v} for k, v in summary_data.items()])
        df_summary.to_excel(writer, sheet_name="Tổng quan", index=False)

    output.seek(0)
    return output.read()


def export_pdf(reviews: list, shop_name: str = "E-ComSight") -> bytes:
    """Export báo cáo PDF chuyên nghiệp"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=20,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1E40AF"),
        spaceAfter=6,
        alignment=TA_CENTER
    )
    heading_style = ParagraphStyle(
        "Heading",
        parent=styles["Heading2"],
        fontSize=13,
        fontName="Helvetica-Bold",
        textColor=colors.HexColor("#1E40AF"),
        spaceBefore=12,
        spaceAfter=6
    )

    elements = []

    # Title
    elements.append(Paragraph("📊 BÁO CÁO PHÂN TÍCH CẢM XÚC KHÁCH HÀNG", title_style))
    elements.append(Paragraph(
        f"{shop_name} · {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        ParagraphStyle("sub", parent=styles["Normal"], alignment=TA_CENTER, textColor=colors.grey)
    ))
    elements.append(HRFlowable(width="100%", thickness=2, color=colors.HexColor("#1E40AF")))
    elements.append(Spacer(1, 12))

    # Summary stats
    df = reviews_to_dataframe(reviews)
    total = len(reviews)
    pos = df["Cảm xúc"].value_counts().get("Tích cực", 0)
    neg = df["Cảm xúc"].value_counts().get("Tiêu cực", 0)
    neu = df["Cảm xúc"].value_counts().get("Trung lập", 0)

    elements.append(Paragraph("I. Tổng quan", heading_style))
    summary_table_data = [
        ["Chỉ số", "Giá trị", "Tỷ lệ"],
        ["Tổng reviews phân tích", str(total), "100%"],
        ["🟢 Tích cực", str(pos), f"{round(pos/total*100,1) if total else 0}%"],
        ["🟡 Trung lập", str(neu), f"{round(neu/total*100,1) if total else 0}%"],
        ["🔴 Tiêu cực", str(neg), f"{round(neg/total*100,1) if total else 0}%"],
    ]
    summary_table = Table(summary_table_data, colWidths=[8*cm, 4*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E2E8F0")),
        ("ROUNDEDCORNERS", [4, 4, 4, 4]),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 16))

    # Top reviews bị cảnh báo
    elements.append(Paragraph("II. Review tiêu cực cần chú ý", heading_style))
    critical_reviews = [r for r in reviews if r.urgency_label in ("critical", "high")][:10]

    if critical_reviews:
        for r in critical_reviews:
            comment_short = (r.comment or "")[:150] + ("..." if len(r.comment or "") > 150 else "")
            urgency_color = "#EF4444" if r.urgency_label == "critical" else "#F97316"
            elements.append(Paragraph(
                f'<font color="{urgency_color}"><b>[{URGENCY_VI.get(r.urgency_label, "").upper()}]</b></font> '
                f'{r.product_name or "N/A"} · ⭐{r.rating_star}',
                styles["Normal"]
            ))
            elements.append(Paragraph(f'"{comment_short}"', styles["Italic"]))
            elements.append(Spacer(1, 8))
    else:
        elements.append(Paragraph("Không có review nghiêm trọng.", styles["Normal"]))

    doc.build(elements)
    output.seek(0)
    return output.read()
